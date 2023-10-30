from numpy import array
from keras.models import Sequential
from keras.layers import LSTM
from keras.layers import Dense
import pandas as pd
import openpyxl
import datetime
from datetime import date
import calendar
open_workbook = 'money_ball.xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
today = date.today()
month = today.strftime("%m/%y") 
year_ne = today.strftime("%Y") 
month_ne= today.strftime("%m")
days = calendar.monthrange(int(year_ne), int(month_ne))[1]
days = int(days)
    

#days = calendar.monthrange(202, month)[1]
def check_empty_row():
    i = 1
    while (ws.cell(row = i, column = 4).value!=None):
        i=i+1
    return i
def get_sum_for_month(month):
    i = check_empty_row()
    sum =0

    for rows in range(2,i):
        if(ws.cell(row = rows, column = 7).value == month):
           if( ws.cell(row= rows, column = 8).value!=None):
                sum = sum + ws.cell(row= rows, column = 8).value
                
                
    return sum

def get_array_for_month(month):
    monthly_expend = []
    i = check_empty_row()
    sum =0

    for rows in range(2,i):
        if(ws.cell(row = rows, column = 7).value == month):
           if( ws.cell(row= rows, column = 8).value!=None):
                sum = sum + ws.cell(row= rows, column = 8).value
                monthly_expend.append( ws.cell(row= rows, column = 8).value )
                
    return monthly_expend
# split a univariate sequence into samples
def split_sequence(sequence, n_steps):
    X, y = list(), list()
    for i in range(len(sequence)):
 # find the end of this pattern
        end_ix = i + n_steps
 # check if we are beyond the sequence
        if end_ix > len(sequence)-(days-n_steps):
            
            break
        seq_x, seq_y = sequence[i:end_ix], sequence[end_ix:end_ix+(days-n_steps)]
        X.append(seq_x)
        if(len(seq_y)==(days-n_steps)):

            y.append(seq_y)
 # gather input and output parts of the pattern
    #seq_x, seq_y = sequence[i:end_ix], sequence[end_ix]   
    return array(X) , array(y)
 
# define input sequence
data = pd.read_csv('dateset.csv', index_col=0)
seq = data['MONEY_PER_DATE']
data = []
for i in range(len(seq)):
    data.append(seq.iloc[i])
inp = get_array_for_month(month)
def sum(yhat):
        sum = 0
        for i in yhat:
            for j in i:
                sum = sum + j
        return sum

def main(data, inp):
    n_steps = len(inp)
    X, y = split_sequence(data, n_steps)
    n_features = 1
    X = X.reshape((X.shape[0], X.shape[1], n_features))
    model = Sequential()
    model.add(LSTM(50, activation='relu', input_shape=(n_steps, n_features)))
    model.add(Dense(days-n_steps))
    model.compile(optimizer='adam', loss='mse')
    model.fit(X, y, epochs=200, verbose=0)
    x_input = array(inp)
    x_input = x_input.reshape((1, n_steps, n_features))
    yhat = model.predict(x_input, verbose=0)
    out_more = get_sum_for_month(month)
    
    su = sum(yhat)+ out_more
    return su

if __name__=="__main__": 
    out = main(data, inp)
    print(out)
    
    