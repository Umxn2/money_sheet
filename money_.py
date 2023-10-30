import openpyxl
import datetime
from datetime import date
import plotext as pl
import os
import subprocess
import lstm
import pandas as pd
import atexit
open_workbook = 'money_ball.xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
ws.cell(row=1, column=1, value = 1)
ws.cell(row=1, column=2, value = 'DATE')
ws.cell(row=1, column=3, value = 'ITEM')
ws.cell(row=1, column=4, value = 'PRICE')
ws.cell(row=1, column=5, value = 'PERSON')
ws.cell(row=1, column=6, value = 'REF_NO')
ws.cell(row=1, column=7, value = 'MONTH')
ws.cell(row=1, column=8, value = 'MONEY_PER_DAY')
ws.cell(row=1, column=9, value = 'MONEY_PER_MONTH')

workbook.save(open_workbook)
today = date.today()
month = today.strftime("%m/%y") 
def exit_handler():
    logs = open("logs.txt", "w")
    today = date.today()
    today =  today.strftime("%d-%b-%Y")
    logs.write(today)
    
    
    


def intitialise_data():
    data = pd.read_csv('dateset.csv', index_col=0)
    seq = data['MONEY_PER_DATE']
    data = []
    for i in range(len(seq)):
        data.append(seq.iloc[i])
    return data
def check_empty_row():
    i = 1
    while (ws.cell(row = i, column = 4).value!=None):
        i=i+1
    return i
def begin():
    i = check_empty_row()
    for rows in range(2,2*i):
        ws.cell(row = rows, column = 1).value=rows
    workbook.save(open_workbook)
    return

def Add_item_on_the_same_day():
    inp_item = input("Hi else what did you buy now?")
    inp_price = int(input("At what price did you buy this?"))
    i = check_empty_row() 
    ws.cell(row=i, column = 3, value = inp_item)
    ws.cell(row=i, column = 4, value = inp_price)
    workbook.save(open_workbook)
    new_out = input("Type 'y' if you would like to add another item else type 'n'")
    if(new_out=='y'):
        Add_item_on_the_same_day()
    else:
        return

def Add_item_on_new_day():
    today = date.today()
    today = today.strftime("%d/%m/%Y")
    today = today.replace("20","")
    i = check_empty_row()
    
    inp_item = input("Hi what did you buy now?")
    inp_price = int(input("At what price did you buy this?"))
    ws.cell(row=i, column = 2, value = today)
    ws.cell(row=i, column = 3, value = inp_item)
    ws.cell(row=i, column = 4, value = inp_price)
    workbook.save(open_workbook)
def clear_workbook():
     for row in ws['A1:G150']:
        for cell in row:
            cell.value = None
            workbook.save(open_workbook)

def remove_row():
    name_removed = input("what item do you want to delete")
    price_removed = int(input("price of the item you want to delete"))
    #date_bought = input("when did you buy this item")
    i = check_empty_row()
    for rows in range(1, i):
        for columns in range(1,5):
                if(ws.cell(row = rows, column = columns).value == price_removed and ws.cell(row = rows, column = columns-1).value == name_removed):
                    print("element found")
                    ws.delete_rows(idx = rows)
                    workbook.save(open_workbook)
                    return
    print("no element like that found")
    return

def print_excel(begin, end):
    i = check_empty_row()
    for rows in range(begin, end):
        for columns in range(1, 9):
            print(f"-|{ws.cell(row = rows, column = columns).value}|-", end ="")
        print(" ")
        
        print(f"{'*'*100}")
def print_excel_for_one_day():
    count =0
    date = input("give date for which you would like to get the excel printed")
   
    i = check_empty_row()
 

    
    for rows in range(1, i):
        
        
        if(ws.cell(row=rows, column = 2).value == date):
            row_number = rows
            count = 1
            break
       
    if(count!=1):
        print("date does not exist")
        return        
    print_excel(row_number, min(row_number+10, i))
    if(min(row_number+10, i)==i):
        return
    print("if you want 10 more values pls enter 'y' else 'n'")
    answer = input()
    if(answer=='y'):
        print_excel(row_number+10, min(row_number+20,i))
    else:
        return
def insertor(ref_no, date, money, person):
    i = check_empty_row()
    for rows in range(2,i):
        if(ws.cell(row=rows, column = 6).value!=ref_no):
            ws.cell(row=i, column = 2, value = date)
            ws.cell(row=i, column = 5, value = person)
            ws.cell(row=i, column = 4, value = money)
            ws.cell(row=i, column = 6, value = ref_no)
            workbook.save(open_workbook)
        else:
            print("already present")
    return
def edit_name():
    switch = 0
    name = input("What item did you buy? ")
    date = input("What date did you buy?")
    person = input("Who did you pay?")
    i = check_empty_row()
    for rows in range(1,i):
        for columns in range(1, 6):
            if(ws.cell(row=rows, column = columns).value==date and ws.cell(row=rows, column= columns+3).value==person):
                ws.cell(row=rows, column= columns+1).value==name
                switch =1
    if(switch==0):
        print("sorry but we couldnt find your item but skill issues happen in life")
    answer = input("Type 'y' if you would like to add more items or type something else to quit")
    if(answer=='y'):
        edit_name()
    else:
        return
def find_with_row_values():
    end = check_empty_row()
    print_excel(1, end)
    row = int(input("What row number item do you wish to change?"))
    columns = input("What parameter do you want to change")
    new = input("what is the new value?")
    columns = columns.lower()
    lis = ["date", "item", "price", "person" ]
    if(columns == lis[0]):
        ws.cell(row=row, column = 2).value = new
    elif(columns == lis[1]):
        ws.cell(row=row, column = 3).value = new
    elif(columns == lis[2]):
        ws.cell(row=row, column = 4).value = new
    elif(columns == lis[3]):
        ws.cell(row=row, column = 5).value = new
    else:
        print("type a suitable value")
    workbook.save(open_workbook)
    answer = input("Do you want to continue, press y for yes else press something else")
    if(answer== 'y'):
        find_with_row_values()
    else:
        return
def find_sum_for_a_day():
    i = check_empty_row()
    answer = input("At what date do you want to get the amount of money you spent?")
    sum = 0
    print(i)
    count = 0
    rows = 1
    while rows!=i:
        

        if(ws.cell(row=rows, column = 2).value == answer):

            count =1
            sum = sum + ws.cell(row=rows, column = 4).value
            rows=rows+1
        if((ws.cell(row=rows, column = 2).value == answer or ws.cell(row=rows, column = 2).value == None) and count==1):
            sum = sum + ws.cell(row=rows, column = 4).value
        else:
            count = 0
        print(sum)
        rows=rows+1

       
    return sum
def sum_of_day():
    i = check_empty_row()
    sum = 0
    for rows in range(2, i):
        if(ws.cell(row = rows, column = 2).value==ws.cell(row = rows+1, column = 2).value):
            sum = sum +  ws.cell(row = rows, column = 4).value
        else:
            sum=sum+ws.cell(row = rows, column = 4).value
            ws.cell(row = rows, column = 8).value=sum
            sum = 0
    workbook.save(open_workbook)
def get_month():
    i = check_empty_row()
    for rows in range(2,i):
        date= ws.cell(row = rows, column = 2).value
        month = date[3:8]
        ws.cell(row = rows, column = 7).value=month
    workbook.save(open_workbook)
def money_per_month():
    i = check_empty_row()
    sum = 0
    for rows in range(2, i):
        if(ws.cell(row = rows, column = 7).value==ws.cell(row = rows+1, column = 7).value):
            sum = sum +  ws.cell(row = rows, column = 4).value
        else:
            sum=sum+ws.cell(row = rows, column = 4).value
            ws.cell(row = rows, column = 9).value=sum
            sum = 0
    workbook.save(open_workbook)
def get_array_for_month(month):
    
    monthly_expend = []
    i = check_empty_row()
    for rows in range(2,i):
        if(ws.cell(row = rows, column = 7).value == month):
           if( ws.cell(row= rows, column = 8).value!=None):
                monthly_expend.append( ws.cell(row= rows, column = 8).value )
    return monthly_expend
                
                
            
   



def statistics(month):
    i = check_empty_row()
    monthly_expend = []
    dates = []
    #month = input("Which month and year would you like")
    sum = 0
    for rows in range(2,i):
        if(ws.cell(row = rows, column = 7).value == month):
           if( ws.cell(row= rows, column = 8).value!=None):
                monthly_expend.append( ws.cell(row= rows, column = 8).value )
                sum = sum + ws.cell(row= rows, column =8).value 
                date = (ws.cell(row= rows, column = 2).value)
                date = date[:2]
                dates.append(date)
   
    print("=========================================")
    print(f"You have spent {sum} this month")
    print("=========================================")
    pl.date_form('d')
    pl.plot(dates, monthly_expend)
    pl.title(f"Money Spent this month = {sum}")
    pl.xlabel("Dates")
    pl.ylabel("Money Spent")
    #pl.colorize("yellow on blue, flash",       "yellow",       "flash",     "blue",          True)
    pl.theme("dark")
    pl.show()
#def store_time():

def get_sum_for_month(month):
    i = check_empty_row()
    sum =0

    for rows in range(2,i):
        if(ws.cell(row = rows, column = 7).value == month):
           if( ws.cell(row= rows, column = 8).value!=None):
                sum = sum + ws.cell(row= rows, column = 8).value
                
                
    return sum



def main():
    print("The program is starting please wait for sometime...")
    
    
    
    
    
    begin()
   
    subprocess.run(["python", './mail_new.py'])  
    
    

    

    


    print("*****************Spend Less*********************")
    print("Press 1 if you want to add item on a new day")
    print("Press 2 if you want to add item on the same day")
    print("Press 3 if you want to remove something from the sheet")
    print("Press 4 if you want to print out excel")
    print("Press 5 if you want to print out excel for one day")
    print("Press 6 if you want to edit out excel")
    print("Press 7 if you want to edit out excel with row and column values ")
    print("Press 8 if you sum of money you spent on some day ")
    print("Press 9 to view month statistics")
    print("Press 10 to predict the amount the money you will spend this month")
    print("Press 15 if you want to exit")
    answer = int(input())
    if answer==1:
        Add_item_on_new_day()
        
        new_out = input("Type 'y' if you would like to add another item else type 'n'")
        if(new_out=='y'):
            Add_item_on_the_same_day()
        else:
            return
        
    if answer==2:
        Add_item_on_the_same_day()
    if answer==3:
        remove_row()
    if answer==4:
        end = check_empty_row()
        print_excel(1, end)
    if answer==5:
        print_excel_for_one_day()
    if answer==6:
        edit_name()
    if answer == 7:
        find_with_row_values()
    if answer == 8:
        sum = find_sum_for_a_day()
        print(sum)
    if answer == 9:
        get_month()
        sum_of_day()
        statistics(month)
    if answer == 10:
        print("The model is running.......")
        
        data = intitialise_data()
        inp = get_array_for_month(month)
        print("It will take sometime please be patient")
        #print(inp)
        output = lstm.main(data, inp)
        sum = get_sum_for_month(month)
        print(f"Amount of money you've spent till now is {sum} ")
        print(f"Amount of money estimated is {output} ")
        #subprocess.run(["python", './lstm.py'])  

    if answer == 15:
        logs = open("logs.txt", "w")
        today = date.today()
        today =  today.strftime("%d-%b-%Y")
        logs.write(today)
        exit()

    
    
    workbook.save(open_workbook)


if __name__=="__main__": 
    main()
    atexit.register(exit_handler)
    

    





