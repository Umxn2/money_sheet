#modules
import imaplib
import email
import html 
import openpyxl

def variable_collector(money, person, date):
    return money, person, date
#credentials
f = open("logs.txt", "r")
date = f.read()

open_workbook = 'money_ball.xlsx'
workbook = openpyxl.load_workbook(open_workbook)
ws = workbook.active
username ="yourmailhere"
app_password= "yourpasswordhere"
gmail_host= 'imap.gmail.com'
mail = imaplib.IMAP4_SSL(gmail_host)
mail.login(username, app_password)
mail.select("INBOX")
_, selected_mails = mail.search(None,f'(SINCE {date})','(FROM "alerts@hdfcbank.net")')
def check_condition(ref_no):
    i = check_empty_row()
    for rows in range(2,i):
        if(ws.cell(row=rows, column = 6).value==ref_no):
            return True
    return False
def check_empty_row():
    i = 1
    while (ws.cell(row = i, column = 4).value!=None):
        i=i+1
    return i
def insertor(ref_no, date, money, person):
    i = check_empty_row()
    for rows in range(2,i):
        if(ws.cell(row=rows, column = 6).value==ref_no):
            print("already present")
            return
            
    ws.cell(row=i, column = 2, value = date)
    ws.cell(row=i, column = 5, value = person)
    ws.cell(row=i, column = 4, value = money)
    ws.cell(row=i, column = 6, value = ref_no)
    workbook.save(open_workbook)
    
    return
for num in selected_mails[0].split():
    _, data = mail.fetch(num , '(RFC822)')
    _, bytes_data = data[0]

    email_message = email.message_from_bytes(bytes_data)
    

    for part in email_message.walk():
        if part.get_content_type()=="text/plain" or part.get_content_type()=="text/html":
            message = part.get_payload(decode=True)
            mess = message.decode()
            
            new_mess = str(mess)
            
            
            start = new_mess.find("Rs.")
            end = new_mess.find("you")
            mess1 = new_mess[start:end]
            
            start = mess1.find("is")
            ref_no = mess1[start+3:start+15]
            #print(ref_no)
            if(check_condition(ref_no)==False):

            

                
                #print(mess1)
                new_end = mess1.find("<br>")
                
                #print(mess1[0:new_end])
                start_money = mess1.find('Rs.')
                finish_money = mess1.find(' has')
                money_var = mess1[start_money+3:finish_money]


                print(money_var)
                start_person = mess1.find('VPA ')
                finish_person = mess1.find(' on')
                person = mess1[start_person+4:finish_person]
                print(person)
                start_date = mess1.find("on ")
                finish_date = mess1.find(". Y")
                date = mess1[start_date+3:finish_date]
                print(date)
                date_new = date.replace("-", "/")
                
                if(money_var!=""):
                    insertor(ref_no, date_new, float(money_var), person)
                

                    print("==========================================")

                break
            
