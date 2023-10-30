import openpyxl
import matplotlib
from matplotlib import pyplot
import numpy as np
from datetime import date
import csv
import os
today = date.today()
month = today.strftime("%m/%y") 
money_ne = today.strftime("%m") 
data_workbook = 'dataset_temp.xlsx'
open_workbook = 'money_ball.xlsx'
final_workbook = 'dataset.xlsx'
workbookbase = openpyxl.load_workbook(open_workbook)
workbookdata = openpyxl.load_workbook(data_workbook)
workbookfinal = openpyxl.load_workbook(final_workbook)
ws = workbookbase.active
ws2 = workbookdata.active
ws3 = workbookfinal.active
ws2.cell(row=1, column=1, value = 1)
ws2.cell(row=1, column=2, value = 'DATE')
ws2.cell(row=1, column=3, value = 'ITEM')
ws2.cell(row=1, column=4, value = 'PRICE')
ws2.cell(row=1, column=5, value = 'PERSON')
ws2.cell(row=1, column=6, value = 'REF_NO')
ws2.cell(row=1, column=7, value = 'MONTH')
ws2.cell(row=1, column=8, value = 'MONEY_PER_DAY')
ws2.cell(row=1, column=9, value = 'MONEY_PER_MONTH')
workbookdata.save(data_workbook)
def check_empty_row2():
    i = 1
    while (ws2.cell(row = i, column = 4).value!=None):
        i=i+1
    return i
def check_empty_row3():
    i = 1
    while (ws3.cell(row = i, column = 4).value!=None):
        i=i+1
    return i
def check_empty_row():
    i = 1
    while (ws.cell(row = i, column = 4).value!=None):
        i=i+1
    return i
def clear_row():
    i = check_empty_row()
    for rows in range(2,i):
        ws.cell(row = rows, column = 6).value=None
    workbookbase.save(open_workbook)
def copy_workbook():
    i = check_empty_row()
    begin=0
    end = 0
    for rows in range(2,i):
        if(ws.cell(row=rows, column = 7).value==month):
            begin=rows
            
            break
    for rows in range(2,i+1):
        if(ws.cell(row=rows, column = 7).value!=month):
            
            end=rows
            
            if(end>begin):
                break        
    
    for rows in range(begin,end):
        ws2.cell(row=rows, column = 2).value = ws.cell(row=rows, column = 2).value
        ws2.cell(row=rows, column = 4).value = ws.cell(row=rows, column = 4).value
        ws2.cell(row=rows, column = 5).value = ws.cell(row=rows, column = 5).value
    workbookdata.save(data_workbook)

def remove_emptyrows():
    for rows in range(1,1700):
        if(ws.cell(row = rows, column = 4).value == None):
            ws.delete_rows(rows)
            print(ws.cell(row = rows, column = 4).value)
            print("deleting row ", rows)
        else:
            continue
    workbookbase.save(open_workbook)
def absurd_entries():
    for rows in range(2,1800):
        if(ws2.cell(row = rows, column = 4).value!=None):
            if(float(ws2.cell(row = rows, column = 4).value) > 1300.0):
                print(ws2.cell(row = rows, column = 4).value)
                ws2.delete_rows(rows)
    workbookdata.save(data_workbook)
        
   

def convert_str_to_float():
    i = check_empty_row()
    for rows in range(2,i):
        ws.cell(row = rows, column = 4).value=float(ws.cell(row = rows, column = 4).value)
    workbookbase.save(open_workbook)
    return

def sum_of_day():
    i = check_empty_row2()
    
    sum = 0
    for rows in range(2, i):
        if(ws2.cell(row = rows, column = 2).value==ws2.cell(row = rows+1, column = 2).value):
            sum = sum +  ws2.cell(row = rows, column = 4).value
        else:
            sum=sum+ws2.cell(row = rows, column = 4).value
            ws2.cell(row = rows, column = 8).value=sum
            sum = 0
    workbookdata.save(data_workbook)

def get_month():
    i = check_empty_row2()
    for rows in range(2,i):
        date= ws2.cell(row = rows, column = 2).value
        month = date[3:8]
        ws2.cell(row = rows, column = 7).value=month
    workbookdata.save(data_workbook)
def money_per_month():
    i = check_empty_row2()
    sum = 0
    for rows in range(2, i):
        if(ws2.cell(row = rows, column = 7).value==ws2.cell(row = rows+1, column = 7).value):
            sum = sum +  ws2.cell(row = rows, column = 4).value
        else:
            sum=sum+ws2.cell(row = rows, column = 4).value
            ws2.cell(row = rows, column = 9).value=sum
            sum = 0
    workbookdata.save(data_workbook)

def grapher():
    i = check_empty_row()
    money_month= []
    month = []
    for rows in range(2,i):
        if(ws.cell(row = rows, column = 9).value!=None):
            
            money_month.append(ws.cell(row = rows, column = 9).value)
            month.append(ws.cell(row = rows, column = 7).value)
    #pyplot.plot(month, money_month)
    return month
    
def graph_per_day():
    i = check_empty_row()
    money_date= []
    date = []
    for rows in range(2,i):
        if(ws.cell(row = rows, column = 8).value!=None):
            
            money_date.append(ws.cell(row = rows, column = 8).value)
            date.append(ws.cell(row = rows, column = 2).value)

    pyplot.plot(date, money_date)
    
    return money_date, date
def clip_data():
    i = check_empty_row()
    for rows in range(2,i):
        if(ws3.cell(row = rows, column = 7).value=="03/22" or ws3.cell(row = rows, column = 7).value=="10/23" ):
            print(ws3.cell(row = rows, column = 7).value)
            ws3.delete_rows(rows)
    workbookfinal.save(final_workbook)
def clear_workbook():
    i = check_empty_row()
    for row in range(2,250):
        for columns in range(1,11):
            
            ws2.cell(row = row, column = columns).value=None
    workbookdata.save(data_workbook)
def get_array_for_month(month):
    
    monthly_expend = []
    date = []
    i = check_empty_row2()
    for rows in range(2,i):
        if(ws2.cell(row = rows, column = 7).value == month):
           if( ws2.cell(row= rows, column = 8).value!=None):
                monthly_expend.append( ws2.cell(row= rows, column = 8).value )
                date.append( ws2.cell(row= rows, column = 2).value)
    return monthly_expend, date

def get_array_for_month_final():
    
    monthly_expend = []
    date = []
    i = check_empty_row3()
    for rows in range(2,i):
        
           if( ws3.cell(row= rows, column = 8).value!=None):
                monthly_expend.append( ws3.cell(row= rows, column = 8).value )
                date.append( ws3.cell(row= rows, column = 2).value)
    
    return monthly_expend, date










# def sum_of_month():
#     i = check_empty_row()
#     sum = 0
#     date =
#     month = ws.cell(row = rows, column = 2).value
    
#     for rows in range(1, i):




        
    
    

             
            
    #



#=remove_emptyrows()
#convert_str_to_float()
#sum_of_day()
#clear_row()




#money, date = graph_per_day()
# #clip_data()
clear_workbook()
copy_workbook()
sum_of_day()
absurd_entries()
get_month()
money_per_month()
money_expend, date = get_array_for_month(month)
money_per_day = money_expend

money_per_day_final, dates_final = get_array_for_month_final()

money_all=money_per_day_final+money_per_day
dates_all = dates_final+date
#print(money_all, dates_all)
os.remove('dateset.csv')

with open('dateset.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    field = ["DATE", "MONEY_PER_DATE"]
    writer.writerow(field)
    for i in range(len(dates_all)):
        writer.writerow([f'{dates_all[i]}', money_all[i]])
    
# #clip_data()
   


